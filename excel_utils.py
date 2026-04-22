from __future__ import annotations

from copy import copy
from datetime import date
from io import BytesIO
from pathlib import Path

import streamlit as st
from openpyxl import Workbook, load_workbook

from address_utils import normalize_text


TARGET_SHEET = "Sheet0"
PRODUCT_SHEET = "Sheet1"


@st.cache_data
def load_product_options(template_path: str) -> list[dict[str, str]]:
    workbook = load_workbook(template_path, data_only=False)
    worksheet = workbook[PRODUCT_SHEET]
    products: list[dict[str, str]] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        product_name = normalize_text(row[0])
        sku_code = normalize_text(row[1])
        if product_name and sku_code:
            products.append({"name": product_name, "sku_code": sku_code})
    return products


def build_output_file(
    template_path: Path,
    address_groups: list[dict[str, object]],
    customer_code: str,
    customer_name: str,
    payment_date: date,
) -> BytesIO:
    workbook = load_workbook(template_path, data_only=False)
    worksheet = workbook[TARGET_SHEET]
    template_row_index = 2

    all_line_items: list[dict[str, object]] = []
    for address_group in address_groups:
        address_info = address_group["address_info"]
        for order_group in address_group["order_groups"]:
            external_order_no = str(order_group["external_order_no"])
            for item in order_group["line_items"]:
                all_line_items.append(
                    {
                        "external_order_no": external_order_no,
                        "receiver": address_info["receiver"],
                        "phone": address_info["phone"],
                        "province": address_info["province"],
                        "city": address_info["city"],
                        "district": address_info["district"],
                        "detail_address": address_info["detail_address"],
                        "product_name": item["product_name"],
                        "sku_code": item["sku_code"],
                        "quantity": item["quantity"],
                    }
                )

    for index, item in enumerate(all_line_items, start=2):
        if index > template_row_index:
            worksheet.insert_rows(index)
            for column in range(1, 17):
                source_cell = worksheet.cell(row=template_row_index, column=column)
                target_cell = worksheet.cell(row=index, column=column)
                if source_cell.data_type == "f" and source_cell.value:
                    target_cell.value = source_cell.value
                elif source_cell.value is not None:
                    target_cell.value = source_cell.value
                target_cell._style = copy(source_cell._style)
                if source_cell.number_format:
                    target_cell.number_format = source_cell.number_format
                if source_cell.font:
                    target_cell.font = copy(source_cell.font)
                if source_cell.fill:
                    target_cell.fill = copy(source_cell.fill)
                if source_cell.border:
                    target_cell.border = copy(source_cell.border)
                if source_cell.alignment:
                    target_cell.alignment = copy(source_cell.alignment)
                if source_cell.protection:
                    target_cell.protection = copy(source_cell.protection)
            worksheet.row_dimensions[index].height = worksheet.row_dimensions[template_row_index].height

        worksheet[f"A{index}"] = str(item["external_order_no"])
        worksheet[f"B{index}"] = payment_date
        worksheet[f"C{index}"] = customer_code
        worksheet[f"D{index}"] = customer_name
        worksheet[f"E{index}"] = ""
        worksheet[f"F{index}"] = "样品"
        worksheet[f"G{index}"] = str(item["receiver"])
        worksheet[f"H{index}"] = str(item["phone"])
        worksheet[f"I{index}"] = ""
        worksheet[f"J{index}"] = str(item["province"])
        worksheet[f"K{index}"] = str(item["city"])
        worksheet[f"L{index}"] = str(item["district"])
        worksheet[f"M{index}"] = str(item["detail_address"])
        worksheet[f"N{index}"] = str(item["sku_code"])
        worksheet[f"O{index}"] = str(item["product_name"])
        worksheet[f"P{index}"] = int(item["quantity"])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def save_output_file(base_dir: Path, output_file: BytesIO, file_name: str) -> Path:
    output_dir = base_dir / "output"
    output_dir.mkdir(exist_ok=True)
    output_path = output_dir / file_name
    output_path.write_bytes(output_file.getvalue())
    return output_path


def build_batch_template_file() -> bytes:
    template_wb = Workbook()
    ws = template_wb.active
    ws.title = "批量地址"
    ws.append(["外部平台单号", "地址"])
    ws.append(["ORDER-001", "张三 13800138000 北京市朝阳区百子湾路 32 号院"])
    ws.append(["ORDER-002", "李四 13900139000 北京市朝阳区百子湾路 33 号院"])
    output = BytesIO()
    template_wb.save(output)
    return output.getvalue()


def parse_uploaded_address_file(uploaded_file) -> list[dict[str, str]]:
    file_name = normalize_text(getattr(uploaded_file, "name", "")).lower()
    rows: list[dict[str, str]] = []

    if not file_name.endswith(".xlsx"):
        raise ValueError("仅支持上传 .xlsx 文件。")

    workbook = load_workbook(BytesIO(uploaded_file.getvalue()), data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    header_row = [normalize_text(cell) for cell in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    for value_row in worksheet.iter_rows(min_row=2, values_only=True):
        row = {
            header_row[index]: normalize_text(value_row[index])
            for index in range(len(header_row))
            if header_row[index]
        }
        if any(row.values()):
            rows.append(row)

    required_headers = {"外部平台单号", "地址"}
    if not rows:
        raise ValueError("上传文件中没有可用数据。")
    missing_headers = [header for header in required_headers if header not in rows[0]]
    if missing_headers:
        raise ValueError(f"上传文件缺少表头：{'、'.join(missing_headers)}")

    normalized_rows = []
    for row in rows:
        external_order_no = normalize_text(row.get("外部平台单号"))
        address_text = normalize_text(row.get("地址"))
        if external_order_no and address_text:
            normalized_rows.append({"external_order_no": external_order_no, "address_text": address_text})

    if not normalized_rows:
        raise ValueError("上传文件中没有同时包含外部平台单号和地址的有效行。")
    return normalized_rows
